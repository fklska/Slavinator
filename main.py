import requests
from telegram import ReplyKeyboardMarkup, Bot, File
from telegram.ext import CommandHandler, MessageHandler, Updater, Filters
from account import Account
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, NamedStyle, Side, Alignment
import os
from dotenv import load_dotenv

load_dotenv()

updater = Updater(token=os.getenv('TOKEN'))


def text_format(text):
    dic = {

    }
    for i in range(len(text.split(';'))):
        dic[i] = text.split(';')[i]
    return dic


def create_account(dic, start_dic):
    numbers = []
    accounts = []
    for value in dic.values():
        numbers.append(value.split(' ')[0])
        numbers.append(value.split(' ')[1])

    numbers = list(set(numbers))

    for value in start_dic.values():
        skip = False

        if value.split(' ')[0] in numbers:
            numbers.remove(value.split(' ')[0])

        for i in accounts:

            if value.split(' ')[0] == i.number:
                i.double = True
                i.opening_balance_credit = int(value.split(' ')[1])
                skip = True

        if skip:
            continue

        if len(value.split(' ')) == 3:
            accounts.append(Account(number=value.split(' ')[0], opening_balance=value.split(' ')[1], status=value.split(' ')[2]))

        else:
            accounts.append(Account(number=value.split(' ')[0], opening_balance=value.split(' ')[1]))

    for number in numbers:
        accounts.append(Account(number))

    return accounts


def account_entires(dic, accounts):
    count = 1
    for wire in dic.values():
        array = wire.split(' ')
        for acc in accounts:
            try:
                if array[0] == acc.number:
                    acc.get_debet(count, array[2])
                if array[1] == acc.number:
                    acc.get_credit(count, array[2])
            except IndexError:
                return 'Введи нормальные данные'
        count += 1
    for acc in accounts:
        acc.get_gross_debet()
        acc.get_gross_credit()
        acc.get_status()
        acc.get_closing_saldo()
    return accounts


def get_codes(update, context):
    chat = update.effective_chat
    create_excel(account_entires(text_format(update.message.text), create_account(text_format(update.message.text), {})))
    context.bot.send_document(
        chat_id=chat.id,
        document=open('sample.xlsx', 'rb'),
    )


def read_excel():
    wb = load_workbook('example.xlsx')
    ws = wb.active
    data = {
        "Start": {

        },
        "Wire": {

        },
        "Sintetic": {

        }
    }

    for row in range(2, 100):
        if ws[f'A{row}'].value is not None:
            data["Start"][row] = ws[f'A{row}'].value

    for row in range(2, 100):
        if ws[f'B{row}'].value is not None:
            data["Wire"][row] = ws[f'B{row}'].value

    for row in range(2, 100):
        if ws[f'C{row}'].value is not None:
            data["Sintetic"][row] = ws[f'C{row}'].value
    return data


def show_message(accounts):
    message = ''
    for i in accounts:
        message += str(i)
    return message


bottom_border = Border(bottom=Side(style='thin'))
right_border = Border(right=Side(style='thin'))
left_border = Border(left=Side(style='thin'))
top_border = Border(top=Side(style='thin'))


def create_excel(datas):
    wb = Workbook()
    ws = wb.active
    col = 1
    row = 3
    for data in datas:
        #print(data)

        ws.merge_cells(f'{ws.cell(row=row-2, column=col).coordinate}:{ws.cell(row=row-2, column=col+1).coordinate}')

        ws.cell(row=row-2, column=col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row-2, column=col).border = bottom_border
        ws.cell(row=row-2, column=col+1).border = bottom_border
        ws.cell(row=row-2, column=col).value = data.number
        ws.cell(row=row-1, column=col).border = Border(right=Side(style='thin'), bottom=Side(style='thin'))
        ws.cell(row=row-1, column=col+1).border = bottom_border

        if data.status == 'Active':
            ws.cell(row=row-1, column=col).value = f'Сн = {data.opening_balance}'
        elif data.status == 'Passive':
            ws.cell(row=row-1, column=col+1).value = f'Сн = {data.opening_balance}'
        elif data.double:
            ws.cell(row=row-1, column=col).value = f'Сн = {data.opening_balance}'
            ws.cell(row=row-1, column=col+1).value = f'Сн = {data.opening_balance_credit}'
        else:
            ws.cell(row=row-1, column=col).value = f'Сн = {data.opening_balance}'
            ws.cell(row=row-1, column=col+1).value = f'Сн = {data.opening_balance}'

        count = 0
        for i in data.debet:

            ws.cell(row=row+count, column=col).border = right_border
            ws.cell(row=row+count, column=col).value = f'{i}){data.debet[i]}'
            count += 1

        count = 0
        for i in data.credit:
            ws.cell(row=row+count, column=col+1).border = left_border
            ws.cell(row=row+count, column=col+1).value = f'{i}){data.credit[i]}'
            count += 1

        if not data.debet and not data.credit:
            ws.cell(row=row, column=col).border = right_border
            ws.cell(row=row+1, column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
            ws.cell(row=row+1, column=col+1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

            ws.cell(row=row+1, column=col).value = f'Обд = {data.gross_debet}'
            ws.cell(row=row+1, column=col+1).value = f'Обк = {data.gross_credit}'

            if data.status == 'Active':
                ws.cell(row=row+2, column=col).value = f'Ск = {data.closing_balance}'
            elif data.status == 'Passive':
                ws.cell(row=row+2, column=col+1).value = f'Ск = {data.closing_balance}'
            else:
                ws.cell(row=row+2, column=col).value = f'Ск = {data.closing_balance_debet}'
                ws.cell(row=row+2, column=col+1).value = f'Ск = {data.closing_balance_credit}'

        else:
            if len(data.debet) > len(data.credit):
                ws.cell(row=row+len(data.debet), column=col).border = right_border
                ws.cell(row=row+len(data.debet), column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                ws.cell(row=row+len(data.debet), column=col+1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

                ws.cell(row=row+len(data.debet), column=col).value = f'Обд = {data.gross_debet}'
                ws.cell(row=row+len(data.debet), column=col+1).value = f'Обк = {data.gross_credit}'

                if data.status == 'Active':
                    ws.cell(row=row+len(data.debet) + 1, column=col).value = f'Ск = {data.closing_balance}'
                elif data.status == 'Passive':
                    ws.cell(row=row+len(data.debet) + 1, column=col+1).value = f'Ск = {data.closing_balance}'
                else:
                    ws.cell(row=row+len(data.debet) + 1, column=col).value = f'Ск = {data.closing_balance_debet}'
                    ws.cell(row=row+len(data.debet) + 1, column=col+1).value = f'Ск = {data.closing_balance_credit}'
            else:
                ws.cell(row=row+len(data.credit), column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                ws.cell(row=row+len(data.credit), column=col+1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

                ws.cell(row=row+len(data.credit), column=col).value = f'Обд = {data.gross_debet}'
                ws.cell(row=row+len(data.credit), column=col+1).value = f'Обк = {data.gross_credit}'

                if data.status == 'Active':
                    ws.cell(row=row+len(data.credit) + 1, column=col).value = f'Ск = {data.closing_balance}'
                elif data.status == 'Passive':
                    ws.cell(row=row+len(data.credit) + 1, column=col+1).value = f'Ск = {data.closing_balance}'
                else:
                    ws.cell(row=row+len(data.credit) + 1, column=col).value = f'Ск = {data.closing_balance_debet}'
                    ws.cell(row=row+len(data.credit) + 1, column=col+1).value = f'Ск = {data.closing_balance_credit}'

        col += 3
        if col > 7:
            row += 10
            col = 1
    last_row = row + 7

    ws.append([])
    ws.merge_cells(f'A{last_row}:A{last_row+1}')
    ws.merge_cells(f'B{last_row}:C{last_row}')
    ws.merge_cells(f'D{last_row}:E{last_row}')
    ws.merge_cells(f'F{last_row}:G{last_row}')
    ws[f'A{last_row}'] = '№'
    ws[f'B{last_row}'] = 'Сальдо начальное'
    ws[f'D{last_row}'] = 'Оборот'
    ws[f'F{last_row}'] = 'Сальдо конечное'

    ws[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'D{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'F{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.append({
        'B': 'Дебет',
        'C': 'Кредит',
        'D': 'Дебет',
        'E': 'Кредит',
        'F': 'Дебет',
        'G': 'Кредит',
        })

    for data in datas:
        if data.status == 'Active':
            ws.append([data.number, data.opening_balance, ' ', data.gross_debet, data.gross_credit, data.closing_balance, ' '])
        elif data.status == 'Passive':
            ws.append([data.number, ' ', data.opening_balance, data.gross_debet, data.gross_credit, ' ', data.closing_balance])
        elif data.double:
            ws.append([data.number, data.opening_balance_credit, data.opening_balance, data.gross_debet, data.gross_credit, data.closing_balance_debet, data.closing_balance_credit])
        else:
            ws.append([data.number, data.opening_balance, data.opening_balance, data.gross_debet, data.gross_credit, data.closing_balance_debet, data.closing_balance_credit])

    ws.append(['', f'=SUM(B{last_row+2}:B{last_row + len(datas)+1})', f'=SUM(C{last_row+2}:C{last_row + len(datas)+1})', f'=SUM(D{last_row+2}:D{last_row + len(datas)+1})', f'=SUM(E{last_row+2}:E{last_row + len(datas)+1})', f'=SUM(F{last_row+2}:F{last_row + len(datas)+1})', f'=SUM(G{last_row+2}:G{last_row + len(datas)+1})'])

    for range in ws[f'A{last_row}:G{last_row + len(datas) + 1}']:
        for article in range:
            article.border = Border(right=Side(style='thin'))

    for range in ws[f'A{last_row+2}:G{last_row+2}']:
        for article in range:
            article.border = Border(top=Side(style='thin'), right=Side(style='thin'))

    for range in ws[f'A{last_row + len(datas) + 1}:G{last_row + len(datas) + 1}']:
        for article in range:
            article.border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

    wb.save("sample.xlsx")


def get_file(update, context):
    chat = update.effective_chat
    file = context.bot.get_file(
        update.message.document.file_id
    )
    File.download(
        file,
        custom_path='example.xlsx')

    context.bot.send_message(
        chat_id=chat.id,
        text='Файл получен',
    )

    print(f'Сообщение отправлено for User: {update.effective_user.username}, В чат: {update.effective_chat.id}')

    create_excel(account_entires(read_excel()['Wire'], create_account(read_excel()['Wire'], read_excel()['Start'])))
    context.bot.send_document(
        chat_id=chat.id,
        document=open('sample.xlsx', 'rb'),
    )


def start(update, context):
    chat = update.effective_chat
    context.bot.send_message(
        chat_id=chat.id,
        text='Привет, это бот который будет помогать тебе делать задачи по бух учету \nТебе нужно отправить данные в excel таблице чтобы бот смог их проанализировать \nВот пример: ',
    )
    context.bot.send_document(
        chat_id=chat.id,
        document=open('template.xlsx', 'rb'),
    )
    context.bot.send_message(
        chat_id=chat.id,
        text="Есть несколько важных особенностей: \n 1) Счета и суммы отделяй пробелом \n 2) Новый счет/новая проводка = новая строчка \n 3) Если Excel сам форматирует числа и удаляет пробелы между ними ставь ' перед данными \n 4) Бот начинает читать файл со 2 строчки \n 5) Про активно-пассивные счета: такие счета считаются сразу для двух сторон как активный, так и пассивный. Если в вашем случае он пассивный значит ваша колонка кредита. Если а-п счет указывается в начальном балансе 2 раза он считается с двумя Сн",
    )


text = "10 20 500;50 51 5500;51 80 5100;70 51 5100;10 20 999;40 41 650;72 73 870"


def main():
    #data = read_excel()
    #print(data['Wire'])
    #print(data['Start'])
    #print(data['Sintetic'])
    #create_excel(account_entires(read_excel()['Wire'], create_account(read_excel()['Wire'], read_excel()['Start'])))

    updater.dispatcher.add_handler(CommandHandler('start', start))
    updater.dispatcher.add_handler(MessageHandler(Filters.document, get_file))
    updater.dispatcher.add_handler(MessageHandler(Filters.text, get_codes))
    updater.start_polling()
    updater.idle()


if __name__ == "__main__":
    main()
