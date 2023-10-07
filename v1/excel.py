from base_classes import Account
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side


class Excel:
    """Class to work and visualize accounting data with excel."""
    def __init__(self) -> None:
        self.bool = False

    def text_format(self, text):
        """Old function to read data from user message."""
        dic = {

        }
        for i in range(len(text.split(';'))):
            dic[i] = text.split(';')[i]
        return dic

    def read_excel(self):
        """Func to read data from xlsx file."""
        wb = load_workbook('v1/excel/example.xlsx')
        ws = wb.active
        data = {
            "Start": {

            },
            "Wire": {

            },
            "Sintetic": {

            }
        }

        # Used to read start data
        # Such as opening balance
        # Number of account
        for row in range(2, 100):
            if ws[f'A{row}'].value is not None:
                data["Start"][row] = ws[f'A{row}'].value

        # Used to read operations data
        # Such as number account of Debet
        # Number account of Creadit
        # And Summ of operation

        for row in range(2, 100):
            if ws[f'B{row}'].value is not None:
                data["Wire"][row] = ws[f'B{row}'].value

        # Used for read Sintetic accounts
        # Not worked
        # Didn't finished it, useless at the moment

        for row in range(2, 100):
            if ws[f'C{row}'].value is not None:
                data["Sintetic"][row] = ws[f'C{row}'].value
        return data

    def create_account(self, dic, start_dic):
        """
        Function to create Account's objects and write start data
        from read info. 2 scenarios when account with opening balance
        and without it.
        """

        numbers = []
        accounts = []
        for value in dic.values():
            numbers.append(value.split(' ')[0])
            numbers.append(value.split(' ')[1])

        numbers = list(set(numbers))

        for value in start_dic.values():
            skip = False

            if value.split(' ')[0] in numbers:
                numbers.remove(value.split(' ')[0])

            for i in accounts:

                if value.split(' ')[0] == i.number:
                    i.double = True
                    i.opening_balance_credit = int(value.split(' ')[1])
                    skip = True

            if skip:
                continue

            if len(value.split(' ')) == 3:
                accounts.append(Account(number=value.split(' ')[0], opening_balance=value.split(' ')[1], status=value.split(' ')[2]))

            else:
                accounts.append(Account(number=value.split(' ')[0], opening_balance=value.split(' ')[1]))

        for number in numbers:
            accounts.append(Account(number))

        return accounts

    def account_entires(self, dic, accounts: [Account]):
        """
        Func to write data in Account objects.
        """

        # This variable used to write operation number
        count = 1

        for wire in dic.values():
            array = wire.split(' ')
            for acc in accounts:
                try:
                    if array[0] == acc.number:
                        acc.get_debet(count, array[2])
                    if array[1] == acc.number:
                        acc.get_credit(count, array[2])
                except IndexError:
                    return 'Введи нормальные данные'
            count += 1
        for acc in accounts:
            acc.get_gross_debet()
            acc.get_gross_credit()
            acc.get_status()
            acc.get_closing_saldo()
        return accounts

    def create_excel(self, datas):
        """
        Function to write processed data in excel
        Borders used to visualize it.
        Several scenarion when account is active/passive/active-passive.
        Very hard to undestand if you didn't know Bookkeeping
        """

        bottom_border = Border(bottom=Side(style='thin'))
        right_border = Border(right=Side(style='thin'))
        left_border = Border(left=Side(style='thin'))
        top_border = Border(top=Side(style='thin'))

        wb = Workbook()
        ws = wb.active
        col = 1
        row = 3
        for data in datas:
            # print(data)

            ws.merge_cells(
                f'{ws.cell(row=row-2, column=col).coordinate}:{ws.cell(row=row-2, column=col+1).coordinate}'
            )

            ws.cell(row=row-2, column=col).alignment = Alignment(
                horizontal='center', vertical='center'
                )
            ws.cell(row=row-2, column=col).border = bottom_border
            ws.cell(row=row-2, column=col+1).border = bottom_border
            ws.cell(row=row-2, column=col).value = data.number
            ws.cell(row=row-1, column=col).border = Border(
                right=Side(style='thin'), bottom=Side(style='thin')
                )
            ws.cell(row=row-1, column=col+1).border = bottom_border

            if data.status == 'Active':
                ws.cell(row=row-1, column=col).value = f'Сн = {data.opening_balance}'

            elif data.status == 'Passive':
                ws.cell(row=row-1, column=col+1).value = f'Сн = {data.opening_balance}'

            elif data.double:
                ws.cell(row=row-1, column=col).value = f'Сн = {data.opening_balance}'
                ws.cell(row=row-1, column=col+1).value = f'Сн = {data.opening_balance_credit}'

            else:
                ws.cell(row=row-1, column=col).value = f'Сн = {data.opening_balance}'
                ws.cell(row=row-1, column=col+1).value = f'Сн = {data.opening_balance}'

            # Write debet data operations
            count = 0
            for i in data.debet:

                ws.cell(row=row+count, column=col).border = right_border
                ws.cell(row=row+count, column=col).value = f'{i}){data.debet[i]}'
                count += 1

            # Write credit data operations
            count = 0
            for i in data.credit:
                ws.cell(row=row+count, column=col+1).border = left_border
                ws.cell(row=row+count, column=col+1).value = f'{i}){data.credit[i]}'
                count += 1

            # Next part of code used to write gross of all operations
            # On debet and credit
            # And closing balance depending on status of account act/pas/act-pas

            # If were not operations close account by Borders
            # Write zero gross balance and closing balance
            if not data.debet and not data.credit:
                ws.cell(row=row, column=col).border = right_border
                ws.cell(row=row+1, column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                ws.cell(row=row+1, column=col+1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

                ws.cell(row=row+1, column=col).value = f'Обд = {data.gross_debet}'
                ws.cell(row=row+1, column=col+1).value = f'Обк = {data.gross_credit}'

                if data.status == 'Active':
                    ws.cell(row=row+2, column=col).value = f'Ск = {data.closing_balance}'
                elif data.status == 'Passive':
                    ws.cell(row=row+2, column=col+1).value = f'Ск = {data.closing_balance}'
                else:
                    ws.cell(row=row+2, column=col).value = f'Ск = {data.closing_balance_debet}'
                    ws.cell(row=row+2, column=col+1).value = f'Ск = {data.closing_balance_credit}'

            # If there are operations
            else:

                # Close account where finish to write
                # Where were most of operations on debet or credit

                if len(data.debet) > len(data.credit):
                    ws.cell(row=row+len(data.debet), column=col).border = right_border
                    ws.cell(row=row+len(data.debet), column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                    ws.cell(row=row+len(data.debet), column=col+1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

                    ws.cell(row=row+len(data.debet), column=col).value = f'Обд = {data.gross_debet}'
                    ws.cell(row=row+len(data.debet), column=col+1).value = f'Обк = {data.gross_credit}'

                    if data.status == 'Active':
                        ws.cell(row=row+len(data.debet) + 1, column=col).value = f'Ск = {data.closing_balance}'
                    elif data.status == 'Passive':
                        ws.cell(row=row+len(data.debet) + 1, column=col+1).value = f'Ск = {data.closing_balance}'
                    else:
                        ws.cell(row=row+len(data.debet) + 1, column=col).value = f'Ск = {data.closing_balance_debet}'
                        ws.cell(row=row+len(data.debet) + 1, column=col+1).value = f'Ск = {data.closing_balance_credit}'

                else:
                    ws.cell(row=row+len(data.credit), column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                    ws.cell(row=row+len(data.credit), column=col+1).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

                    ws.cell(row=row+len(data.credit), column=col).value = f'Обд = {data.gross_debet}'
                    ws.cell(row=row+len(data.credit), column=col+1).value = f'Обк = {data.gross_credit}'

                    if data.status == 'Active':
                        ws.cell(row=row+len(data.credit) + 1, column=col).value = f'Ск = {data.closing_balance}'
                    elif data.status == 'Passive':
                        ws.cell(row=row+len(data.credit) + 1, column=col+1).value = f'Ск = {data.closing_balance}'
                    else:
                        ws.cell(row=row+len(data.credit) + 1, column=col).value = f'Ск = {data.closing_balance_debet}'
                        ws.cell(row=row+len(data.credit) + 1, column=col+1).value = f'Ск = {data.closing_balance_credit}'

            # Base gap through with write Accounts
            # 3 Account per line with a gap of 3 cells between each other
            # every 10 lines
            col += 3
            if col > 7:
                row += 10
                col = 1

        # Next part of code used to write
        # Summarizing information on all accounts on the last row

        last_row = row + 7

        ws.append([])
        ws.merge_cells(f'A{last_row}:A{last_row+1}')
        ws.merge_cells(f'B{last_row}:C{last_row}')
        ws.merge_cells(f'D{last_row}:E{last_row}')
        ws.merge_cells(f'F{last_row}:G{last_row}')
        ws[f'A{last_row}'] = '№'
        ws[f'B{last_row}'] = 'Сальдо начальное'
        ws[f'D{last_row}'] = 'Оборот'
        ws[f'F{last_row}'] = 'Сальдо конечное'

        ws[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'F{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.append({
            'B': 'Дебет',
            'C': 'Кредит',
            'D': 'Дебет',
            'E': 'Кредит',
            'F': 'Дебет',
            'G': 'Кредит',
            })

        for data in datas:
            if data.status == 'Active':
                ws.append([data.number, data.opening_balance, ' ', data.gross_debet, data.gross_credit, data.closing_balance, ' '])
            elif data.status == 'Passive':
                ws.append([data.number, ' ', data.opening_balance, data.gross_debet, data.gross_credit, ' ', data.closing_balance])
            elif data.double:
                ws.append([data.number, data.opening_balance_credit, data.opening_balance, data.gross_debet, data.gross_credit, data.closing_balance_debet, data.closing_balance_credit])
            else:
                ws.append([data.number, data.opening_balance, data.opening_balance, data.gross_debet, data.gross_credit, data.closing_balance_debet, data.closing_balance_credit])

        ws.append(['', f'=SUM(B{last_row+2}:B{last_row + len(datas)+1})', f'=SUM(C{last_row+2}:C{last_row + len(datas)+1})', f'=SUM(D{last_row+2}:D{last_row + len(datas)+1})', f'=SUM(E{last_row+2}:E{last_row + len(datas)+1})', f'=SUM(F{last_row+2}:F{last_row + len(datas)+1})', f'=SUM(G{last_row+2}:G{last_row + len(datas)+1})'])

        for range in ws[f'A{last_row}:G{last_row + len(datas) + 1}']:
            for article in range:
                article.border = Border(right=Side(style='thin'))

        for range in ws[f'A{last_row+2}:G{last_row+2}']:
            for article in range:
                article.border = Border(top=Side(style='thin'), right=Side(style='thin'))

        for range in ws[f'A{last_row + len(datas) + 1}:G{last_row + len(datas) + 1}']:
            for article in range:
                article.border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

        wb.save("v1/excel/sample.xlsx")
